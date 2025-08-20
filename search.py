import requests
from openpyxl import Workbook, load_workbook
import os
import schedule
import time
from datetime import datetime, timedelta

# Masukkan API Key Anda dari https://newsapi.org
API_KEY = "6acfa02130c541bbac83ce7babb999ff"
BASE_URL = "https://newsapi.org/v2/everything"
FILENAME = "hasil_berita_harian.xlsx"

KEYWORDS = ["oknum marinir", "oknum tni al", "oknum tni", "oknum aparat"]  # daftar keyword

def cari_berita(keyword, jumlah=10):
    waktu_sekarang = datetime.utcnow()
    waktu_24jam = waktu_sekarang - timedelta(hours=24)

    params = {
        "q": keyword,
        "apiKey": API_KEY,
        "language": "id",
        "sortBy": "publishedAt",
        "pageSize": jumlah,
        "from": waktu_24jam.strftime("%Y-%m-%dT%H:%M:%S"),
        "to": waktu_sekarang.strftime("%Y-%m-%dT%H:%M:%S")
    }
    response = requests.get(BASE_URL, params=params)
    data = response.json()

    if data.get("status") != "ok":
        print(f"Gagal ambil data untuk '{keyword}':", data)
        return []

    hasil = []
    for artikel in data["articles"]:
        hasil.append([
            artikel["title"],          # Judul
            artikel["source"]["name"], # Sumber
            artikel["url"],            # Link
            artikel["publishedAt"]     # Tanggal
        ])
    return hasil

def simpan_excel(keyword, berita, tanggal_sheet):
    if os.path.exists(FILENAME):
        wb = load_workbook(FILENAME)
    else:
        wb = Workbook()
        wb.remove(wb.active)

    sheet_name = (keyword[:20] + "_" + tanggal_sheet)[:31]

    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        ws = wb.create_sheet(title=sheet_name)
        ws.append(["Judul", "Sumber", "URL", "Tanggal"])

    for row in berita:
        ws.append(row)

    wb.save(FILENAME)

def job():
    waktu = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    tanggal_sheet = datetime.now().strftime("%Y-%m-%d")

    print(f"\n⏳ Update berita otomatis ({waktu})")

    for keyword in KEYWORDS:
        berita = cari_berita(keyword, jumlah=10)
        if not berita:
            continue

        # === tampilkan hasil di terminal ===
        print(f"\n📌 Keyword: {keyword} (24 jam terakhir, {len(berita)} berita)\n")
        for i, artikel in enumerate(berita, 1):
            judul, sumber, url, tanggal = artikel
            print(f"{i}. {judul}")
            print(f"   Sumber : {sumber}")
            print(f"   Link   : {url}")
            print(f"   Tanggal: {tanggal}\n")

        # simpan ke Excel
        simpan_excel(keyword, berita, tanggal_sheet)

    print(f"✅ Data berita tersimpan ke {FILENAME}\n")

if __name__ == "__main__":
    # Jalankan tiap 1 menit
    schedule.every(1).minutes.do(job)

    print("Scheduler aktif ✅ (update berita setiap menit, hanya 24 jam terakhir)")
    while True:
        schedule.run_pending()
        time.sleep(30)
